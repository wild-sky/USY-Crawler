import requests
import execjs
import re
import openpyxl as xl


class JWXT():
    def __init__(self, userAccount, userPassword):
        '''
        教务系统类，填写完参数后将自动调用Login \n
        :param userAccount: 登录账号
        :param userPassword: 登录密码
        '''
        self.userAccount = userAccount
        self.userPassword = userPassword
        self.loginCookies = None
        self.request = requests.Session()
        self.name = ''
        with open('lock.js', 'r') as file:
            content = file.read()
        self.js = execjs.compile(content)
        self.__Login()

    def getSecretContent(self, s):
        '''
        得到客户端的加密字符串
        :param s: 欲加密文本
        :return: 返回加密字符串
        '''
        s = str(s)
        return self.js.eval('encodeInp("{param}")'.format(param=s))

    def __enconded(self):
        '''
        对账户密码进行编码，作为Login的post参数
        :return: 返回一个包含用来作为Login的post参数的字典
        '''
        account = self.getSecretContent(self.userAccount)
        pwd = self.getSecretContent(self.userPassword)
        return {
            'userAccount': account,
            'userPassword': pwd,
            'encoded': account + '%%%' + pwd
        }

    def __Login(self):
        '''
        登录
        :return: 登录后的Response
        '''
        url = 'http://jwxt.sanyau.edu.cn/syxy_jsxsd/xk/LoginToXk'
        res = self.request.post(url, data=self.__enconded())
        res.encoding = 'utf-8'
        tip = re.findall('<font.*?>(.*)</font>', res.text)
        if tip != []:
            print('提示：', tip[0])
            exit(100)
        self.name = re.findall('<p style="font-weight: 500;">(.*?)</p>', res.text)[0]
        print(self.name, "登录成功！")
        return res.text

    def getScore(self, term='2020-2021-1'):
        '''
        得到指定学期的成绩
        :param term: 查询的学期
        :return: 返回一个成绩列表，第0个为标题
        '''
        url = 'http://jwxt.sanyau.edu.cn/syxy_jsxsd/kscj/cjcx_list?kksj=' + term
        res = self.request.get(url)
        res.encoding = 'utf-8'
        trArray = re.findall('''<tr>([\s\S]*?)</tr>''', res.text)
        title = re.findall('''<th.*?>([\s\S]*?)</th>''', trArray[0])
        for i in range(len(title)):
            title[i] = title[i].replace("""\r\n\t\t\t """, "")
        r = [title]
        for tr in trArray[1:]:
            subject = re.findall('<td.*?>([\s\S]*?)</td', tr)
            subject[4] = re.findall('([\u4e00-\u9fa5]|\d.*)', subject[4])[0].replace('\r', '')
            r.append(subject)
        return r

    @staticmethod
    def getGPA(scoreArr):
        '''
        :param scoreArr: 通过JWXT.getScore()函数获取
        :return: 绩点
        '''
        sumValue = .0  # 学分总和
        sum = .0  # 学分*绩点总和
        for subject in scoreArr[1:]:
            power = float(subject[8])  # 绩点，即权重
            if power == 0:
                continue
            value = float(subject[6])  # 学分
            sum += value * power
            sumValue += value
        return sum / sumValue


if __name__ == '__main__':
    userName = '****'  # 账号
    passWord = '****'  # 密码
    jwxt = JWXT(userName, passWord)  # 登录教务系统
    res = jwxt.getScore()  # 拿到成绩Arr
    # 将成绩写入excel导出
    wb = xl.Workbook()  # 创建一个Excel
    ws = wb.active
    for row in range(len(res)):
        for col in range(len(res[row])):
            ws.cell(row + 1, col + 1, res[row][col])
    GPA = jwxt.getGPA(res)
    ws.cell(1, len(res[0]) + 1, "绩点")
    ws.cell(2, len(res[0]) + 1, GPA)
    wb.save('成绩单.xlsx')
    wb.close()
    print('done')
